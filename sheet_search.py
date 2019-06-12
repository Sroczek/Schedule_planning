from openpyxl import load_workbook
from collections import defaultdict
import intervals as I
import datetime

begin_classes_time = datetime.time(hour=8, minute=0)
end_classes_time = datetime.time(hour=19, minute=20)

day_map = {'Pn': 0, 'Wt': 1, 'Sr': 2, 'Cz': 3, 'Pt': 4, 'Sb': 5, 'Nd': 6}

cm = {
    'bud': 'bud',  # Nazwa budynku w zeszycie "sale"
    'nr': 'nr',  # Nazwa sali w zeszycie "sale"
    'typ': 'typ',  # Typ sali w zeszycie "sale"
    'dzien': 'dzien', # Dzien zajec w zeszytach zajec
    'godz': 'godz',  # Godzina zajec w zeszytach zajec
    'koniec': 'koniec',  # Koniec zajec w zeszytach zajec
    'osoba': 'osoba',  # Osoba prowadzaca zajecia
    'studia': 'studia',  # Rodzaj studiow
    'sem': 'sem',  # Nr semestru
    'sala': 'sala',
    'przedmiot': 'przedmiot',  # Nazwa przedmiotu w zeszytach zajec
    'zima_s': 'zima-s',  # Zeszyt zawierajacy zajecia zimowe stacjonarne
    'zima_n': 'zima_n',  # Zeszyt zawierajacy zajecia zimowe niestacjonarne
    'lato_s': 'lato_s',  # Zeszyt zawierajacy zajecia letnie stacjonarne
    'lato_n': 'lato_n',  # Zeszyt zawierajacy zajecia letnie niestacjonarne
    'zima_inne': 'zima_inne',  # Zeszyt zawierajacy zajecia zimowe inne
    'lato_inne': 'lato_inne',  # Zeszyt zawierajacy zajecia letnie inne
    'sale': 'sale'  # Zeszyt zawierajacy sale w ktorych moga odbywac sie zajecia
}


def dtoi(d: datetime.time):
    return d.hour * 60 + d.minute


def itod(i: int):
    return datetime.time(hour=i // 60, minute=i % 60).strftime("%H:%M")


def blocking_row_to_interval(br):
    if br[0] is None or br[1] is None:
        return []
    if br[2] is None:
        return [br[0], I.closedopen(dtoi(br[1]), dtoi(br[1]) + 90)]
    return [br[0], I.closedopen(dtoi(br[1]), dtoi(br[2]))]


class SheetSearch:

    def __rows_generator(self, ws):
        for i, row in enumerate(ws.iter_rows(min_row=2, max_col=16)):
            yield [i + 2] + [cell.value for cell in row]

    def __get_titles(self, ws):
        for row in ws.iter_rows(min_row=1, max_row=1, max_col=16):
            return ["numer"] + [cell.value for cell in row]

    def __map_data(self, ws):
        titles = self.__get_titles(ws)
        for row in self.__rows_generator(ws):
            yield {titles[i]: cell for i, cell in enumerate(row)}

    def __map_rooms(self, ws):
        gen = self.__map_data(ws)
        for row in gen:
            if row[cm["bud"]] == "D17" or row[cm["bud"]] == "D8":
                yield {"nazwa": ((row[cm["bud"]] if row[cm["bud"]] is not None else "") + " " + (
                    row[cm["nr"]] if row[cm["nr"]] is not None else "")).strip(),
                       "typ": row[cm["typ"]] if row[cm["typ"]] is not None else ""}

    def get_blocking_rows(self, ws, row_no):
        """Metoda, ktora zwraca wszystkie terminy, w ktorych prowadzacy lub dany rocznik jest zajety"""
        magic_row = list(filter(lambda x: x["numer"] == row_no, list(self.rdd_dict[ws])))[0]
        return list(map(lambda x: blocking_row_to_interval((x[cm["dzien"]], x[cm["godz"]], x[cm["koniec"]])),
                    filter(lambda x: x[cm["osoba"]] == magic_row[cm["osoba"]] or (x[cm["studia"]] == magic_row[cm["studia"]]
                                                                          and x[cm["sem"]] ==magic_row[cm["sem"]]), self.rdd_dict[ws])))

    def get_proper_rooms(self, ws, row_no):
        """Metoda, ktora zwraca nazwy wszystkich sal, w ktorych moga odbyc sie dane zajecia"""
        magic_row = list(filter(lambda x: x["numer"] == row_no, self.rdd_dict[ws]))[0]
        return list(
            set(map(lambda x: x["nazwa"], filter(lambda x: x[cm["typ"]] == magic_row[cm["typ"]], self.rdd_dict[cm["sale"]]))))

    def get_blocking_rows_for_rooms(self, ws, room_list):
        """Metoda, ktora zwraca slownik przypisujacy kazdej sali trojki: dzien,godzina,koniec,
         w ktorych odbywaja sie zajecia"""
        help = list(set(map(lambda x: (x[cm["sala"]], x[cm["dzien"]], x[cm["godz"]], x[cm["koniec"]]),
                            filter(lambda x: x[cm["sala"]] in room_list, self.rdd_dict[ws]))))
        result = defaultdict(list)
        for x in room_list:
            result[x] = []
        for i, j, k, l in help:
            result[i].append(blocking_row_to_interval((j, k, l)))
        return result

    def get_row(self, ws, row_no):
        magic_row = list(filter(lambda x: x["numer"] == row_no, list(self.rdd_dict[ws])))[0]
        return magic_row

    def find_possible_hours(self, ws, row_no):
        row = self.get_row(ws, row_no)
        class_duration = dtoi(row[cm["koniec"]]) - dtoi(row[cm["godz"]]) if row[cm["koniec"]] is not None else 90

        print("Wyszukiwanie: {}, {}, {}, {}".format(row[cm["przedmiot"]], row[cm["osoba"]], row[cm["sala"]], row[cm["godz"]]))

        if ws[-1] == 's':
            days = ('Pn', 'Wt', 'Sr', 'Cz', 'Pt')
        elif ws[-1] == 'n':
            days = ('Sb', 'Nd')

        br = self.get_blocking_rows(ws, row_no)

        days_unreserved_hours = {day: I.closed(dtoi(begin_classes_time), dtoi(end_classes_time)) for day in days}

        for interval in br:
            if interval:
                days_unreserved_hours[interval[0]] = days_unreserved_hours[interval[0]] - interval[1]

        pr = self.get_proper_rooms(ws, row_no)
        brr = self.get_blocking_rows_for_rooms(ws, pr)
        #brr2 = self.get_blocking_rows_for_rooms(ws[:-1] + "inne", pr)
        brr2 = self.get_blocking_rows_for_rooms(cm['zima_inne'], pr) if ws == cm['zima_n'] or ws == cm['zima_s'] \
            else self.get_blocking_rows_for_rooms(cm['lato_inne'], pr)
        for key, value in brr.items():
            for key2, value2 in brr2.items():
                if key == key2:
                    value += value2
        result = []

        for room, intervals_when_blocked in brr.items():

            days_unreserved_room_hours = {day: I.closed(dtoi(begin_classes_time), dtoi(end_classes_time)) for day in
                                          days}

            for interval in intervals_when_blocked:
                if interval and interval[0] in days:
                    days_unreserved_room_hours[interval[0]] = days_unreserved_room_hours[interval[0]] - interval[1]

            for day in days:
                possible_hours = days_unreserved_hours[day] & days_unreserved_room_hours[day]
                result += [(room, day, interv) for interv in list(possible_hours)]

        result = list(filter(lambda y: y[2].upper - y[2].lower >= class_duration, result))
        result = sorted(result, key=lambda x: x[2].lower)
        result = sorted(result, key=lambda x: day_map[x[1]])

        return result

    def __init__(self, path):
        wb = load_workbook(filename=path, read_only=True)
        self.rdd_dict = {
            cm["zima_s"]: [x for x in self.__map_data(wb[cm["zima_s"]])],
            cm["lato_s"]: [x for x in self.__map_data(wb[cm["lato_s"]])],
            cm["zima_n"]: [x for x in self.__map_data(wb[cm["zima_n"]])],
            cm["lato_n"]: [x for x in self.__map_data(wb[cm["lato_n"]])],
            cm["zima_inne"]: [x for x in self.__map_data(wb[cm["zima_inne"]])],
            cm["lato_inne"]: [x for x in self.__map_data(wb[cm["lato_inne"]])],
            cm["sale"]: [x for x in self.__map_rooms(wb[cm["sale"]])]
        }


# Przyklad uzycia
#
# if __name__ == '__main__':
#     ss = SheetSearch("sheet3.xlsx")
#     br = ss.get_blocking_rows("zima_s", 10)
#     pr = ss.get_proper_rooms("zima_s", 10)
#     brr = ss.get_blocking_rows_for_rooms("zima_inne", pr)
#     brr2 = ss.get_blocking_rows_for_rooms("zima_s", pr)
#     print(ss.get_row("zima_s",10))
#     print("Zajete terminy:")
#     for item in br:
#         print(item)
#     print('\n')
#     print("Zajetosc sal:")
#     for key, value in brr.items():
#         for key2, value2 in brr2.items():
#             if key == key2:
#                 value += value2
#     for key, value in brr.items():
#         print(key)
#         for v in value:
#             print(v)
